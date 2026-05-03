"""
Content extraction module - extracts text from various file types
"""
from pathlib import Path
from typing import Optional
import mimetypes

from utils.logger import setup_logger
from config import config

logger = setup_logger('extractor')

class ContentExtractor:
    """Extracts text content from various file types"""
    
    def __init__(self):
        self.preview_length = config.CONTENT_PREVIEW_LENGTH
    
    def extract_content(self, file_path: Path) -> str:
        """
        Extract text content from file
        
        Args:
            file_path: Path to file
            
        Returns:
            Extracted text content (preview)
        """
        extension = file_path.suffix.lower()
        
        try:
            # Route to appropriate extractor
            if extension in ['.txt', '.md', '.log', '.csv', '.json', '.xml', '.yaml', '.yml']:
                return self._extract_text(file_path)
            
            elif extension in ['.py', '.js', '.jsx', '.ts', '.tsx', '.html', '.css', 
                              '.java', '.cpp', '.c', '.h', '.go', '.rs', '.rb', '.php']:
                return self._extract_code(file_path)
            
            elif extension == '.pdf':
                return self._extract_pdf(file_path)
            
            elif extension in ['.doc', '.docx']:
                return self._extract_word(file_path)
            
            elif extension in ['.xls', '.xlsx']:
                return self._extract_excel(file_path)
            
            elif extension in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.svg']:
                return self._extract_image_metadata(file_path)
            
            else:
                logger.debug(f"No extractor for {extension}")
                return ""
                
        except Exception as e:
            logger.debug(f"Error extracting content from {file_path.name}: {e}")
            return ""
    
    def _extract_text(self, file_path: Path) -> str:
        """Extract from plain text files"""
        try:
            # Try UTF-8 first
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read(self.preview_length * 2)  # Read a bit more
                return content[:self.preview_length]
        except Exception as e:
            logger.debug(f"Error reading text file {file_path.name}: {e}")
            return ""
    
    def _extract_code(self, file_path: Path) -> str:
        """Extract from code files (same as text but might add more logic later)"""
        return self._extract_text(file_path)
    
    def _extract_pdf(self, file_path: Path) -> str:
        """Extract text from PDF files"""
        try:
            import PyPDF2
            
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                
                # Extract first few pages
                text_parts = []
                max_pages = min(3, len(reader.pages))  # First 3 pages max
                
                for page_num in range(max_pages):
                    page = reader.pages[page_num]
                    text_parts.append(page.extract_text())
                
                full_text = ' '.join(text_parts)
                return full_text[:self.preview_length]
                
        except ImportError:
            logger.warning("PyPDF2 not installed, skipping PDF extraction")
            return ""
        except Exception as e:
            logger.debug(f"Error extracting PDF {file_path.name}: {e}")
            return ""
    
    def _extract_word(self, file_path: Path) -> str:
        """Extract text from Word documents"""
        try:
            import docx
            
            doc = docx.Document(file_path)
            
            # Extract paragraphs
            text_parts = []
            for para in doc.paragraphs:
                text_parts.append(para.text)
                
                # Stop if we have enough
                current_length = sum(len(t) for t in text_parts)
                if current_length >= self.preview_length:
                    break
            
            full_text = ' '.join(text_parts)
            return full_text[:self.preview_length]
            
        except ImportError:
            logger.warning("python-docx not installed, skipping Word extraction")
            return ""
        except Exception as e:
            logger.debug(f"Error extracting Word doc {file_path.name}: {e}")
            return ""
    
    def _extract_excel(self, file_path: Path) -> str:
        """Extract text from Excel files"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Get first sheet
            sheet = wb.active
            
            # Extract cell values
            text_parts = []
            max_rows = min(50, sheet.max_row)  # First 50 rows
            
            for row in sheet.iter_rows(max_row=max_rows, values_only=True):
                row_text = ' '.join(str(cell) for cell in row if cell is not None)
                text_parts.append(row_text)
                
                # Stop if we have enough
                current_length = sum(len(t) for t in text_parts)
                if current_length >= self.preview_length:
                    break
            
            wb.close()
            
            full_text = ' '.join(text_parts)
            return full_text[:self.preview_length]
            
        except ImportError:
            logger.warning("openpyxl not installed, skipping Excel extraction")
            return ""
        except Exception as e:
            logger.debug(f"Error extracting Excel {file_path.name}: {e}")
            return ""
    
    def _extract_image_metadata(self, file_path: Path) -> str:
        """Extract metadata from images (no OCR for now)"""
        try:
            from PIL import Image
            
            with Image.open(file_path) as img:
                # Get basic info
                info_parts = [
                    f"Image: {img.format}",
                    f"Size: {img.size[0]}x{img.size[1]}",
                    f"Mode: {img.mode}"
                ]
                
                # Get EXIF data if available
                exif = img.getexif()
                if exif:
                    # Common EXIF tags
                    interesting_tags = {
                        271: "Make",
                        272: "Model", 
                        306: "DateTime",
                        270: "ImageDescription"
                    }
                    
                    for tag_id, name in interesting_tags.items():
                        if tag_id in exif:
                            info_parts.append(f"{name}: {exif[tag_id]}")
                
                return ' '.join(info_parts)
                
        except ImportError:
            logger.warning("Pillow not installed, skipping image metadata")
            return ""
        except Exception as e:
            logger.debug(f"Error extracting image metadata from {file_path.name}: {e}")
            return ""
    
    def get_searchable_text(self, file_path: Path, file_name: str) -> str:
        """
        Get complete searchable text for a file
        
        Args:
            file_path: Path to file
            file_name: Name of file
            
        Returns:
            Combined searchable text (filename + content)
        """
        # Start with filename (important for search)
        parts = [file_name]
        
        # Add file name without extension
        name_without_ext = Path(file_name).stem
        if name_without_ext != file_name:
            parts.append(name_without_ext)
        
        # Extract content
        content = self.extract_content(file_path)
        if content:
            parts.append(content)
        
        # Join with spaces
        return ' '.join(parts)

# Convenience function
def extract_content(file_path: Path) -> str:
    """Extract content from file"""
    extractor = ContentExtractor()
    return extractor.extract_content(file_path)

if __name__ == "__main__":
    # Test extractor
    import sys
    from .scanner import FileScanner
    
    if len(sys.argv) > 1:
        test_path = sys.argv[1]
    else:
        test_path = "."
    
    print(f"\n{'='*60}")
    print(f"Testing Content Extractor")
    print(f"{'='*60}\n")
    
    # Scan for files
    scanner = FileScanner()
    files = list(scanner.scan_directory(test_path))[:10]  # First 10 files
    
    # Extract content
    extractor = ContentExtractor()
    
    for file_info in files:
        print(f"\nFile: {file_info.name}")
        print(f"Extension: {file_info.extension}")
        
        content = extractor.extract_content(file_info.path)
        if content:
            preview = content[:100] + "..." if len(content) > 100 else content
            print(f"Content preview: {preview}")
        else:
            print("No content extracted")
        print("-" * 60)